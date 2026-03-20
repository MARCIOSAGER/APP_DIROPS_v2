"""Generate a preview Excel showing how AIAAN data will be imported into DIROPS."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict
import os

CITY_TO_ICAO = {
    'CABINDA': 'FNCA', 'CATUMBELA': 'FNBG', 'LUBANGO': 'FNUB', 'SAURIMO': 'FNSA',
    'SOYO': 'FNSO', 'HUAMBO': 'FNHU', 'ONDJIVA': 'FNGI', 'LUENA': 'FNUE',
    'MENONGUE': 'FNME', 'DUNDO': 'FNDU', 'KUITO': 'FNKU', 'LUANDA': 'FNLU',
    'JOHANNESBURG': 'FAOR', 'JOHANESBURGO': 'FAOR', 'CAPE TOWN': 'FACT', 'CAPETOWN': 'FACT',
    'LAGOS': 'DNMM', 'MAPUTO': 'FQMA', 'ADDIS ABEBA': 'HAAB', 'NAIROBI': 'HKJK',
    'WINDHOEK': 'FYWH', 'KINSHASA': 'FZAA', 'BRAZZAVILLE': 'FCBB', 'LISBOA': 'LPPT',
    'PARIS': 'LFPG', 'FRANKFURT': 'EDDF', 'DUBAI': 'OMDB', 'DOHA': 'OTHH',
    'BOM JESUS': 'FNBJ',
}
OPERATOR_TO_ICAO = {
    'TAAG': 'DT', 'TAP AIR PORTUGAL': 'TP', 'ETHIOPIAN AIRLINES': 'ET',
    'EMIRATES': 'EK', 'AIR FRANCE': 'AF', 'LUFTHANSA': 'LH', 'QATAR': 'QR',
    'ASKY': 'KP', 'TURKISH AIRLINES': 'TK', 'ROYAL AIR MAROC': 'AT',
    'HELIMALONGO': 'HM', 'HELI MALONGO': 'HM',
}

def is_empty(v):
    if v is None: return True
    s = str(v).strip()
    return s in ('', '-', ' -', '- ')

def parse_time(val):
    if val is None: return None
    if hasattr(val, 'strftime'): return val.strftime('%H:%M')
    if isinstance(val, (int, float)):
        m = int(val * 24 * 60)
        return f'{m // 60:02d}:{m % 60:02d}'
    s = str(val).strip()
    if s in ('', '-', ' -'): return None
    parts = s.split(':')
    if len(parts) >= 2:
        try: return f'{int(parts[0]):02d}:{int(parts[1]):02d}'
        except: pass
    return None

def norm_reg(r):
    if not r: return ''
    return str(r).strip().upper().replace('-', '').replace(' ', '')

# -------------------------------------------------------------------
base = os.path.join(os.path.dirname(__file__), '..')
wb = openpyxl.load_workbook(os.path.join(base, 'docs', 'AIAAN VOOS 2025.xlsx'), read_only=True, data_only=True)
ws = wb['AIAAN VOOS 2025']

parsed = []
for row in ws.iter_rows(min_row=2, max_row=201, values_only=True):
    data = row[0]
    if not data: continue
    ata, std, atd = row[5], row[6], row[7]
    has_ata = not is_empty(ata)
    has_std = not is_empty(std)
    if has_ata and not has_std: tipo = 'ARR'
    elif has_std and not has_ata: tipo = 'DEP'
    else: continue

    date_str = data.strftime('%Y-%m-%d') if hasattr(data, 'strftime') else str(data)[:10]
    callsign = str(row[10]).strip() if row[10] else ''
    reg = norm_reg(row[11])
    op = str(row[4]).strip().upper() if row[4] else ''
    dest = str(row[15]).strip().upper() if row[15] else ''
    pmd = row[12]
    pax_emb = int(row[16]) if row[16] and not is_empty(row[16]) else 0
    pax_des = int(row[18]) if row[18] and not is_empty(row[18]) else 0
    transit = int(row[17]) if row[17] and not is_empty(row[17]) else 0
    crew = int(row[19]) if row[19] and not is_empty(row[19]) else 0
    carga = float(row[25]) if row[25] and not is_empty(row[25]) else 0
    stand = str(row[9]).strip() if row[9] else ''
    manga_on = parse_time(row[30]) if len(row) > 30 else None
    pca_on = parse_time(row[32]) if len(row) > 32 else None
    gpu_on = parse_time(row[36]) if len(row) > 36 else None

    parsed.append({
        'data': date_str, 'tipo': tipo, 'callsign': callsign, 'reg': reg,
        'op': op, 'op_icao': OPERATOR_TO_ICAO.get(op, '?'),
        'dest': dest, 'dest_icao': CITY_TO_ICAO.get(dest, '?'),
        'pmd_ton': pmd, 'ata': parse_time(ata), 'std': parse_time(std), 'atd': parse_time(atd),
        'pax_emb': pax_emb, 'pax_des': pax_des, 'transit': transit, 'crew': crew,
        'carga': carga, 'stand': stand,
        'manga': manga_on, 'pca': pca_on, 'gpu': gpu_on,
    })

# Find pairs
groups = defaultdict(list)
for i, r in enumerate(parsed):
    groups[(r['reg'], r['data'])].append(i)

pairs = []
for key, indices in groups.items():
    arrs = sorted([i for i in indices if parsed[i]['tipo'] == 'ARR'], key=lambda i: parsed[i].get('ata') or '00:00')
    deps = sorted([i for i in indices if parsed[i]['tipo'] == 'DEP'], key=lambda i: parsed[i].get('std') or '00:00')
    used = set()
    for a in arrs:
        for d in deps:
            if d in used: continue
            if (parsed[d].get('std') or '23:59') >= (parsed[a].get('ata') or '00:00'):
                pairs.append((a, d))
                used.add(d)
                break

# Create output
out = openpyxl.Workbook()
hfill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
hfont = Font(color='FFFFFF', bold=True, size=10)
arr_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
dep_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
pair_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')

# Sheet 1: Voos Parseados
ws1 = out.active
ws1.title = 'Voos Parseados'
h1 = ['Data', 'Tipo', 'Callsign', 'Registo', 'Operador', 'ICAO Op.', 'Destino/Origem', 'ICAO Dest.', 'PMD(t)', 'ATA', 'STD', 'ATD', 'Pax Emb', 'Pax Des', 'Transit', 'Crew', 'Carga(kg)', 'Stand', 'PBB', 'PCA', 'GPU']
for c, h in enumerate(h1, 1):
    cell = ws1.cell(row=1, column=c, value=h)
    cell.font = hfont; cell.fill = hfill; cell.alignment = Alignment(horizontal='center')

for ri, r in enumerate(parsed, 2):
    fill = arr_fill if r['tipo'] == 'ARR' else dep_fill
    vals = [r['data'], r['tipo'], r['callsign'], r['reg'], r['op'], r['op_icao'], r['dest'], r['dest_icao'], r['pmd_ton'], r['ata'], r['std'], r['atd'], r['pax_emb'], r['pax_des'], r['transit'], r['crew'], r['carga'], r['stand'], r['manga'], r['pca'], r['gpu']]
    for c, v in enumerate(vals, 1):
        cell = ws1.cell(row=ri, column=c, value=v)
        cell.fill = fill

# Sheet 2: Pares Linkados
ws2 = out.create_sheet('Pares ARR-DEP')
h2 = ['Registo', 'PMD(t)', 'Data', 'Voo ARR', 'ATA', 'Voo DEP', 'STD', 'ATD', 'Estac.(min)', 'Pax Emb', 'Carga(kg)', 'Tipo', 'Origem (ARR)', 'Destino (DEP)', 'PBB', 'PCA', 'GPU']
for c, h in enumerate(h2, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = hfont; cell.fill = hfill; cell.alignment = Alignment(horizontal='center')

for pi, (a, d) in enumerate(pairs, 2):
    ar, dr = parsed[a], parsed[d]
    ata_t = ar.get('ata') or '00:00'
    atd_t = dr.get('atd') or dr.get('std') or '00:00'
    try:
        h1_, m1 = map(int, ata_t.split(':'))
        h2_, m2 = map(int, atd_t.split(':'))
        perm = (h2_ * 60 + m2) - (h1_ * 60 + m1)
        if perm < 0: perm += 1440
    except: perm = 0
    dest_icao = CITY_TO_ICAO.get(ar['dest'], ar['dest'])
    is_intl = not dest_icao.startswith('FN') if dest_icao and dest_icao != '?' else False

    vals = [ar['reg'], ar['pmd_ton'], ar['data'], ar['callsign'], ar['ata'], dr['callsign'], dr['std'], dr['atd'], perm, dr['pax_emb'], dr['carga'], 'INT' if is_intl else 'DOM', ar['dest'], dr['dest'], ar['manga'] or dr['manga'], ar['pca'] or dr['pca'], ar['gpu'] or dr['gpu']]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=pi, column=c, value=v)
        if pi % 2 == 0: cell.fill = pair_fill

# Sheet 3: Resumo
ws3 = out.create_sheet('Resumo')
ws3.cell(row=1, column=1, value='Preview Importacao AIAAN -> DIROPS').font = Font(bold=True, size=14)
ws3.cell(row=3, column=1, value='Total voos parseados:').font = Font(bold=True)
ws3.cell(row=3, column=2, value=len(parsed))
ws3.cell(row=4, column=1, value='Chegadas (ARR):').font = Font(bold=True)
ws3.cell(row=4, column=2, value=sum(1 for r in parsed if r['tipo'] == 'ARR'))
ws3.cell(row=5, column=1, value='Partidas (DEP):').font = Font(bold=True)
ws3.cell(row=5, column=2, value=sum(1 for r in parsed if r['tipo'] == 'DEP'))
ws3.cell(row=6, column=1, value='Pares linkados:').font = Font(bold=True)
ws3.cell(row=6, column=2, value=len(pairs))
ws3.cell(row=7, column=1, value='Aeroporto:').font = Font(bold=True)
ws3.cell(row=7, column=2, value='FNBJ')
ws3.cell(row=9, column=1, value='Amostra: primeiras 200 linhas do Excel AIAAN').font = Font(italic=True, color='888888')
ws3.cell(row=10, column=1, value='Verde = ARR (Chegada) | Laranja = DEP (Partida)').font = Font(italic=True, color='888888')

# Adjust widths
for ws_obj in [ws1, ws2, ws3]:
    for col in ws_obj.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_obj.column_dimensions[col[0].column_letter].width = min(max_len + 3, 25)

outpath = os.path.join(base, 'docs', 'AIAAN_preview_importacao.xlsx')
out.save(outpath)
print(f'Ficheiro criado: {outpath}')
print(f'  - {len(parsed)} voos parseados')
print(f'  - {len(pairs)} pares linkados')
